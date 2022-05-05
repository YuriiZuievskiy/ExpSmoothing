import openpyxl
import numpy
import matplotlib.pyplot as plt
from sklearn.metrics import r2_score


wookbook = openpyxl.load_workbook("Dataset_augmented.xlsx") # Define variable to load the wookbook
worksheet = wookbook.active # Define variable to read the active sheet:

def read_column(int): #Iterate the loop to read the cell values and return list of values in column.
 ls = []
 for i in range(1, worksheet.max_row):
  for col in worksheet.iter_cols(int, int):
   ls.append(col[i].value)
 return ls

power = int(input("Введіть степінь полінома:"))

x = read_column(1)
i = read_column(2)
j = read_column(3)
k = read_column(4)
ask1  = input("Показати значення стовпчиків в таблиці? Y/N:")
if ask1=="Y":
    print ("\n",
        "Чистая прибыль:",x,"\n",
        "Исследования и разработки",i,"\n",
        "Общие операционные расходы",j,"\n",
        "денежные средства от опер. Деятельности",k)
elif ask1!="Y":
    pass

mymodel = numpy.poly1d(numpy.polyfit(x, i, power))
mymodel1 = numpy.poly1d(numpy.polyfit(x, j, power))
mymodel2 = numpy.poly1d(numpy.polyfit(x, k, power))
myline = numpy.linspace(1, 100, 19)

#Calculating accuracy of regression:
print ('Math square error for model 1:', r2_score(i, mymodel(x)))
print ('Math square error for model 2:', r2_score(j, mymodel1(x)))
print ('Math square error for model 3:', r2_score(k, mymodel2(x)))
print ('Math square root error for model 1:', numpy.sqrt(r2_score(i, mymodel(x))))
print ('Math square root error for model 2:', numpy.sqrt(r2_score(j, mymodel1(x))))
print ('Math square root error for model 3:', numpy.sqrt(r2_score(k, mymodel2(x))))

#Building plot
plt.scatter(x, i)
plt.scatter(x, j)
plt.scatter(x, k)
plt.style.use('Solarize_Light2')
ax = plt.subplot(111)
ax.plot(x, i, label='Чистая прибыль')
ax.plot(x, j, label='Исследования и разработки')
ax.plot(x, k, label='денежные средства от опер. Деятельности')
plt.plot(myline, mymodel(myline))
plt.plot(myline, mymodel1(myline))
plt.plot(myline, mymodel2(myline))
ax.plot(myline, mymodel(myline), label='Model Чистая прибыль')
ax.plot(myline, mymodel1(myline), label='Model Исследования и разработки')
ax.plot(myline, mymodel2(myline), label='Model денежные средства от опер. Деятельности')
if power == 1:
    plt.title('Linear Regression')
else:
    plt.title('Polinomial Regression')
chartBox = ax.get_position()
ax.set_position([chartBox.x0, chartBox.y0, chartBox.width*0.6, chartBox.height])
ax.legend(loc='upper center', bbox_to_anchor=(1.45, 0.8), shadow=True, ncol=1)

plt.show()

"""print ('Math absolute error:', metrics.mean_absolute_error(i,mymodel(myline)))
print ('Math square error:', metrics.mean_squared_error(j,mymodel1(myline)))
print ('Math square root error:', numpy.sqrt(metrics.mean_squared_error(k,mymodel2(myline))))"""

ask2  = input("Показати значення отриманої регресії? Y/N:")
if ask2=="Y":
    print ("Resulting Regression Data\n",
       "Исследования и разработки",mymodel(myline),"\n",
       "Общие операционные расходы",mymodel1(myline),"\n",
       "денежные средства от опер. Деятельности",mymodel2(myline))
elif ask2!="Y":
    pass